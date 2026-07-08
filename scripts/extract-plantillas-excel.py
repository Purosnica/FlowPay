"""Extract message templates from CREDICOMPRAS Excel plantillas sheet."""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print('Install openpyxl: pip install openpyxl', file=sys.stderr)
    sys.exit(1)

EXCEL_PATH = Path(
    r'C:\Users\Bryan Silva\Downloads\CARTERA CREDITO PARA COMPRAS 16032026 (1BLOQ).xlsm'
)
OUT_PATH = (
    Path(__file__).parent.parent / 'prisma' / 'data' / 'plantillas-credicompras.json'
)

ETAPA_MAP = {
    'SMS ILOCALIZADOS': ('SMS', 'ADMINISTRATIVA'),
    'SMS PROMOCION DE DESCUENTO': ('SMS', 'ADMINISTRATIVA'),
    'OFERTA DE DESCUENTO': ('WHATSAPP', 'ADMINISTRATIVA'),
    'PNC': ('WHATSAPP', 'EXTRAJUDICIAL'),
}


def normalizar_texto(val: object) -> str | None:
    if val is None:
        return None
    text = str(val).strip()
    if not text:
        return None
    return re.sub(r'\n{3,}', '\n\n', text)


def parametrizar(contenido: str) -> str:
    """Replace sample debtor data with template variables."""
    out = contenido
    out = re.sub(
        r'\*KARLA CECILIA BAQUEDANO FERNANDEZ\*',
        '*{{nombre}}*',
        out,
        flags=re.I,
    )
    out = re.sub(
        r'KARLA CECILIA BAQUEDANO FERNANDEZ',
        '{{nombre}}',
        out,
        flags=re.I,
    )
    out = re.sub(r'\*353451\*', '*{{prestamo}}*', out)
    out = re.sub(r'\*615-150471-0000T\*', '*{{documento}}*', out)
    out = re.sub(
        r'C\$\s*[\d,]+\.?\d*',
        '{{saldo}}',
        out,
    )
    out = re.sub(
        r'\d{2}/\d{2}/\d{4}',
        '{{fechaLimite}}',
        out,
    )
    out = re.sub(
        r'TICTAC S\.A\.?',
        '{{mandanteLegal}}',
        out,
        flags=re.I,
    )
    out = re.sub(
        r'Crédito para Compras|CREDICOMPRAS',
        '{{mandante}}',
        out,
        flags=re.I,
    )
    return out

def inferir_canal_etapa(nombre: str) -> tuple[str, str]:
    for key, (canal, etapa) in ETAPA_MAP.items():
        if key.lower() in nombre.lower():
            return canal, etapa
    upper = nombre.upper()
    if 'SMS' in upper:
        return 'SMS', 'ADMINISTRATIVA'
    if 'DESPACHO' in upper or 'LEGAL' in upper or 'PNC' in upper:
        return 'WHATSAPP', 'EXTRAJUDICIAL'
    if 'DESCUENTO' in upper or 'FERIA' in upper:
        return 'WHATSAPP', 'ADMINISTRATIVA'
    return 'WHATSAPP', 'ADMINISTRATIVA'


def extraer_desde_hoja_plantillas(ws) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = 0
    for i, row in enumerate(rows[:5]):
        for cell in row:
            if cell and 'SMS ILOCALIZADOS' in str(cell).upper():
                header_idx = i
                break

    header_row = rows[header_idx]
    templates: list[dict[str, str]] = []

    for col_idx, header in enumerate(header_row):
        if header is None:
            continue
        nombre = str(header).strip()
        if len(nombre) < 3 or nombre.isdigit():
            continue

        partes: list[str] = []
        for row in rows[header_idx + 1 :]:
            if col_idx >= len(row):
                continue
            texto = normalizar_texto(row[col_idx])
            if texto:
                partes.append(texto)

        if not partes:
            continue

        contenido = parametrizar('\n\n'.join(partes))
        canal, etapa = inferir_canal_etapa(nombre)

        templates.append({
            'nombre': nombre.strip(),
            'canal': canal,
            'etapa': etapa,
            'contenido': contenido,
        })

    return templates


def main() -> None:
    if not EXCEL_PATH.exists():
        print(f'Excel not found: {EXCEL_PATH}', file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(
        EXCEL_PATH,
        read_only=True,
        data_only=True,
        keep_vba=True,
    )

    templates: list[dict[str, str]] = []
    if 'plantillas' in wb.sheetnames:
        templates = extraer_desde_hoja_plantillas(wb['plantillas'])
        print(f'Plantillas sheet: {len(templates)} column templates')

    wb.close()

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(templates, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    print(f'Wrote {len(templates)} templates to {OUT_PATH}')
    for t in templates:
        print(f"  - {t['nombre']} ({t['canal']}, {t['etapa']}) — {len(t['contenido'])} chars")


if __name__ == '__main__':
    main()
